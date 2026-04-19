#!/usr/bin/env python3
"""Import historical design requests from Excel into Supabase.
Handles Excel tables where activity_name / activity_period are blank on sub-rows
(forward-fill from the last non-blank row above).
"""

import warnings
warnings.filterwarnings('ignore')

import openpyxl
import requests
import json
from datetime import datetime, date

EXCEL_PATH = "/Volumes/MU P100 1TB/claude_司/設計需求系統/淨淨2026 設計需求表xlsx.xlsx"
SUPABASE_URL = "https://hhebbtnkpjjmyoiqtidd.supabase.co"
ANON_KEY = "sb_publishable_N0mVM4buKlxlRzmWe-_zSw_ftZvrmih"

HEADERS = {
    "apikey": ANON_KEY,
    "Authorization": f"Bearer {ANON_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=minimal",
}

CHANNEL_MAP = {
    "官網-1月活動":     "website-activity",
    "官網-2月活動":     "website-activity",
    "官網-3月活動":     "website-activity",
    "官網-4月活動":     "website-activity",
    "官網-5月活動":     "website-activity",
    "門市旗艦店":       "flagship-store",
    "百貨-快閃櫃":      "department-popup",
    "百貨館內宣傳輸出":  "department-instore",
    "蝦皮活動需求":     "shopee",
    "MO店+活動需求":    "mo-store",
    "Line@專區":        "line-at",
    "FB廣告圖文":       "fb-ads",
    "團購":             "group-buy",
    "粉專需求":         "facebook-page",
    "CRM需求":          "crm",
    "CRM-VIP需求":      "crm-vip",
    "戶外廣告":         "outdoor-ads",
    "異業合作":         "cross-industry",
    "小瑄需求":         "distributor",
}

STATUS_MAP = {
    "完成": "completed", "已完成": "completed",
    "進行中": "in_progress",
    "待處理": "pending",
    "待審核": "review",
    "需修改": "revision",
}

BATCH_SIZE = 50


def get_channels():
    resp = requests.get(f"{SUPABASE_URL}/rest/v1/channels?select=id,slug", headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return {ch["slug"]: ch["id"] for ch in resp.json()}


def to_date_str(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def to_str(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def to_int(value, default=1):
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def insert_batch(records):
    resp = requests.post(
        f"{SUPABASE_URL}/rest/v1/design_requests",
        headers=HEADERS,
        data=json.dumps(records, ensure_ascii=False).encode("utf-8"),
        timeout=60,
    )
    if resp.status_code not in (200, 201):
        raise RuntimeError(f"Insert failed {resp.status_code}: {resp.text[:500]}")
    return len(records)


def main():
    print("Fetching channels...")
    slug_to_id = get_channels()
    print(f"  {len(slug_to_id)} channels found\n")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    total_inserted = 0
    total_skipped = 0

    for sheet_name in wb.sheetnames:
        slug = CHANNEL_MAP.get(sheet_name)
        if not slug:
            print(f"[SKIP] {sheet_name} — no mapping")
            continue

        channel_id = slug_to_id.get(slug)
        if not channel_id:
            print(f"[SKIP] {sheet_name} — slug '{slug}' not in DB")
            continue

        ws = wb[sheet_name]
        batch = []
        sheet_inserted = 0
        sheet_skipped = 0

        # Forward-fill state
        last_activity_name = None
        last_activity_period = None
        last_designer = None
        last_requester = None
        last_deadline = None
        last_status_raw = None

        for row in ws.iter_rows(min_row=2, values_only=True):
            col_a = to_str(row[0] if len(row) > 0 else None)   # 活動期間
            col_b = to_str(row[1] if len(row) > 1 else None)   # 活動內容
            col_c = to_str(row[2] if len(row) > 2 else None)   # 用途
            col_d = to_str(row[3] if len(row) > 3 else None)   # 尺寸
            col_e = row[4] if len(row) > 4 else None            # 數量
            col_f = to_str(row[5] if len(row) > 5 else None)   # 文案
            col_g = to_str(row[6] if len(row) > 6 else None)   # 需求說明
            col_h = row[7] if len(row) > 7 else None            # 截止日
            col_i = to_str(row[8] if len(row) > 8 else None)   # 負責人
            col_j = to_str(row[9] if len(row) > 9 else None)   # 進度
            col_k = to_str(row[10] if len(row) > 10 else None) # 發需求的人

            # A row is "real data" only if it has purpose (col_c) or size_spec (col_d)
            # or a fresh activity_name (col_b). Rows with none of these are truly empty.
            has_content = bool(col_b or col_c or col_d)
            if not has_content:
                sheet_skipped += 1
                # Reset forward-fill on completely empty rows
                last_activity_name = None
                last_activity_period = None
                last_designer = None
                last_requester = None
                last_deadline = None
                last_status_raw = None
                continue

            # Update forward-fill values when present
            if col_b:
                last_activity_name = col_b
            if col_a:
                last_activity_period = col_a
            if col_i:
                last_designer = col_i
            if col_k:
                last_requester = col_k
            if col_h:
                last_deadline = col_h
            if col_j:
                last_status_raw = col_j

            # Must have an activity_name (from this row or forward-filled)
            activity_name = last_activity_name
            purpose = col_c

            if not activity_name:
                sheet_skipped += 1
                continue

            # Use forward-filled values for deadline/designer/requester
            deadline = to_date_str(last_deadline)
            designer_name = last_designer
            requester_name = last_requester
            status_raw = last_status_raw
            status = STATUS_MAP.get(status_raw or "", "completed")

            parts = ["歷史匯入"]
            if designer_name:
                parts.append(f"設計師:{designer_name}")
            if requester_name:
                parts.append(f"發案:{requester_name}")

            record = {
                "channel_id":      channel_id,
                "activity_period": last_activity_period,
                "activity_name":   activity_name,
                "purpose":         purpose or "設計需求",
                "size_spec":       col_d,
                "quantity":        to_int(col_e),
                "copywriting":     col_f,
                "product_info":    col_g,
                "deadline":        deadline,
                "status":          status,
                "priority":        "normal",
                "notes":           " | ".join(parts),
                "requester_id":    None,
                "designer_id":     None,
            }
            batch.append(record)

            if len(batch) >= BATCH_SIZE:
                try:
                    sheet_inserted += insert_batch(batch)
                    total_inserted += len(batch)
                    batch = []
                except RuntimeError as e:
                    print(f"  [ERROR] {e}")
                    batch = []

        if batch:
            try:
                sheet_inserted += insert_batch(batch)
                total_inserted += len(batch)
            except RuntimeError as e:
                print(f"  [ERROR] final batch: {e}")

        total_skipped += sheet_skipped
        print(f"  {sheet_name:20s} → {sheet_inserted:3d} 筆  (跳過 {sheet_skipped} 空行)")

    print(f"\n完成：匯入 {total_inserted} 筆，跳過 {total_skipped} 空行")


if __name__ == "__main__":
    main()
